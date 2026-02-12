# app.py - 合并后的完整文件
import os
import json
import re
from datetime import datetime, date, timedelta
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
from flask_wtf import FlaskForm
from wtforms import StringField, SubmitField, PasswordField, SelectField
from wtforms.validators import DataRequired, Length
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import io
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import zipfile
import requests
import tempfile
import shutil
import urllib
from openai import OpenAI




app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY') or 'your-secret-key-here'
app.config['DATA_DIR'] = 'data'  # JSON文件存储目录
app.config['ADMIN_ACCOUNTS_FILE'] = os.path.join(app.config['DATA_DIR'], 'admin_accounts.json')
MAX_DAILY_REQUESTS = 50 
app.config['DOWNLOAD_TIMEOUT'] = 30  # 下载超时时间（秒）
app.config['MAX_CONCURRENT_DOWNLOADS'] = 5  # 最大并发下载数
app.config['SONG_DOWNLOAD_DIR'] = os.path.join(app.config['DATA_DIR'], 'downloads')
os.makedirs(app.config['SONG_DOWNLOAD_DIR'], exist_ok=True)
app.config['DEEPSEEK_API_KEY'] = os.environ.get('DEEPSEEK_API_KEY') or '替换为你的deepseekapi'


# 确保数据目录存在
os.makedirs(app.config['DATA_DIR'], exist_ok=True)
import sys

# 定义年级和班级选项
GRADE_CLASS_OPTIONS = {
    '初一': [f'初一{i}班' for i in range(1, 19)],  # 1-18班
    '初二': [f'初二{i}班' for i in range(1, 11)],  # 1-10班
    '初三': [f'初三{i}班' for i in range(1, 11)],  # 1-10班
    '高一': [f'高一{i}班' for i in range(1, 13)],  # 1-12班
    '高二': [f'高二{i}班' for i in range(1, 11)],  # 1-10班
    '高三': [f'高三{i}班' for i in range(1, 11)]   # 1-10班
}

# 系统状态管理
STATUS_FILE = os.path.join(app.config['DATA_DIR'], 'system_status.json')

def get_system_status():
    """获取系统状态"""
    if os.path.exists(STATUS_FILE):
        try:
            with open(STATUS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    
    # 默认状态
    return {
        'requests_paused': False,
        'pause_reason': ''
    }

def save_system_status(status):
    """保存系统状态"""
    try:
        with open(STATUS_FILE, 'w', encoding='utf-8') as f:
            json.dump(status, f, ensure_ascii=False, indent=2)
        return True
    except:
        return False

def is_requests_paused():
    """检查点歌是否被暂停"""
    status = get_system_status()
    return status.get('requests_paused', False)

# 安全辅助函数
def sanitize_input(input_string, max_length=None):
    """
    清理用户输入，防止XSS攻击
    移除HTML标签和JavaScript代码，并限制长度
    """
    if not input_string:
        return input_string
    
    # 移除HTML标签
    cleaned = re.sub(r'<[^>]*>', '', input_string)
    # 移除JavaScript事件处理程序
    cleaned = re.sub(r'on\w+=\s*["\']?[^"\']*["\']?', '', cleaned)
    # 移除其他可能危险的属性
    cleaned = re.sub(r'(href|src|data)=\s*["\']?javascript:[^"\']*["\']?', '', cleaned)
    
    # 限制长度
    if max_length and len(cleaned) > max_length:
        cleaned = cleaned[:max_length]
    
    return cleaned

def validate_date_string(date_string):
    """
    验证日期字符串格式是否为YYYY-MM-DD
    """
    try:
        datetime.strptime(date_string, '%Y-%m-%d')
        return True
    except ValueError:
        return False

# 表单类 - 点歌表单
class SongRequestForm(FlaskForm):
    song_name = StringField('歌曲名称', validators=[
        DataRequired(message='歌曲名称不能为空'),
        Length(min=1, max=100, message='歌曲名称长度不能超过100个字符')
    ])
    
    grade = SelectField('年段', choices=[
        ('', '请选择年段'),
        ('初一', '初一'),
        ('初二', '初二'),
        ('初三', '初三'),
        ('高一', '高一'),
        ('高二', '高二'),
        ('高三', '高三')
    ], validators=[DataRequired(message='请选择年段')])
    
    class_name = SelectField('班级', choices=[('', '请先选择年段')], validators=[DataRequired(message='请选择班级')])
    
    student_name = StringField('姓名', validators=[
        DataRequired(message='姓名不能为空'),
        Length(min=2, max=4, message='姓名长度必须在2-4个字符之间')
    ])
    
    submit = SubmitField('提交')

# 表单类 - 登录表单
class LoginForm(FlaskForm):
    username = StringField('用户名', validators=[
        DataRequired(message='用户名不能为空')
    ])
    
    password = PasswordField('密码', validators=[
        DataRequired(message='密码不能为空')
    ])
    
    submit = SubmitField('登录')

# JSON文件处理函数
def get_today_filename(create_if_not_exists=False):
    """获取当天数据文件的路径"""
    # 判断当前时间是否在18:00之后
    now = datetime.now()
    target_date = date.today()
    
    # 如果当前时间在下午18:00之后，则使用明天的日期
    if now.hour >= 18:
        target_date = target_date + timedelta(days=1)
    
    today_str = target_date.isoformat()
    filename = os.path.join(app.config['DATA_DIR'], f"{today_str}.json")
    
    # 只有在明确要求时才创建文件
    if create_if_not_exists and not os.path.exists(filename):
        save_daily_list([])
    
    return filename

def get_daily_list(date_str=None):
    """获取指定日期的歌曲列表"""
    if not date_str:
        filename = get_today_filename()
    else:
        # 防止路径遍历攻击
        if not validate_date_string(date_str):
            return []
        filename = os.path.join(app.config['DATA_DIR'], f"{date_str}.json")
    
    # 确保文件路径在数据目录内
    if not os.path.abspath(filename).startswith(os.path.abspath(app.config['DATA_DIR'])):
        return []
    
    if os.path.exists(filename):
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return []
    return []

def save_daily_list(data, date_str=None):
    """保存指定日期的歌曲列表"""
    if not date_str:
        filename = get_today_filename()
    else:
        # 防止路径遍历攻击
        if not validate_date_string(date_str):
            return False
        filename = os.path.join(app.config['DATA_DIR'], f"{date_str}.json")
    
    # 确保文件路径在数据目录内
    if not os.path.abspath(filename).startswith(os.path.abspath(app.config['DATA_DIR'])):
        return False
    
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except:
        return False
    
def get_daily_request_count():
    """获取当日已点歌曲数量"""
    today_list = get_daily_list()
    return len(today_list)

def get_remaining_requests():
    """获取剩余可点歌数量"""
    count = get_daily_request_count()
    return max(0, MAX_DAILY_REQUESTS - count)
def download_single_song(song_id, song_name, artist=''):
    """
    下载单首歌曲到本地
    """
    if not song_id:
        return None, "缺少歌曲ID"
    
    try:
        # 使用API获取歌曲下载链接
        params = {
            'url': song_id,
            'level': 'standard',
            'type': 'json'
        }
        
        response = requests.get(
            'https://api.zh-mc.top/Song_V1',
            params=params,
            timeout=app.config['DOWNLOAD_TIMEOUT']
        )
        response.raise_for_status()
        
        song_data = response.json()
        if not song_data.get('success') or not song_data.get('data'):
            return None, song_data.get('message', '获取歌曲信息失败')
        
        download_url = song_data['data'].get('url')
        if not download_url:
            return None, "无法获取歌曲下载链接"
        
        # 下载歌曲
        song_response = requests.get(
            download_url, 
            stream=True,
            timeout=app.config['DOWNLOAD_TIMEOUT'],
            headers={
                'Referer': 'http://music.126.net/',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36'
            }
        )
        song_response.raise_for_status()
        
        def sanitize_filename(name):
            """统一的文件名清理函数"""
            # 移除或替换非法字符
            illegal_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
            for char in illegal_chars:
                name = name.replace(char, '_')
            
            # 移除控制字符
            name = ''.join(char for char in name if ord(char) >= 32)
            
            # 限制长度
            if len(name) > 100:
                name = name[:100]
            
            return name.strip()
        
        # 生成文件名
        filename_title = song_name
        song_filename = f"{sanitize_filename(filename_title)}.mp3"
        song_filepath = os.path.join(app.config['SONG_DOWNLOAD_DIR'], song_filename)
        
        # 保存歌曲文件
        with open(song_filepath, 'wb') as f:
            for chunk in song_response.iter_content(chunk_size=8192):
                f.write(chunk)
        
        return song_filepath, "下载成功"
    
    except requests.exceptions.RequestException as e:
        return None, f"网络错误: {str(e)}"
    except Exception as e:
        return None, f"下载失败: {str(e)}"

def download_single_song(song_id, song_name, artist=''):
    """
    下载单首歌曲到本地，并返回歌词信息
    """
    if not song_id:
        return None, "缺少歌曲ID", None
    
    try:
        # 使用API获取歌曲下载链接
        params = {
            'url': song_id,
            'level': 'standard',
            'type': 'json'
        }
        
        response = requests.get(
            'https://api.zh-mc.top/Song_V1',
            params=params,
            timeout=app.config['DOWNLOAD_TIMEOUT']
        )
        response.raise_for_status()
        
        song_data = response.json()
        if not song_data.get('success') or not song_data.get('data'):
            return None, song_data.get('message', '获取歌曲信息失败'), None
        
        download_url = song_data['data'].get('url')
        lyric = song_data['data'].get('lyric', '')  # 获取歌词
        
        if not download_url:
            return None, "无法获取歌曲下载链接", lyric
        
        # 下载歌曲
        song_response = requests.get(
            download_url, 
            stream=True,
            timeout=app.config['DOWNLOAD_TIMEOUT'],
            headers={
                'Referer': 'http://music.126.net/',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36'
            }
        )
        song_response.raise_for_status()
        
        def sanitize_filename(name):
            """统一的文件名清理函数"""
            # 移除或替换非法字符
            illegal_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
            for char in illegal_chars:
                name = name.replace(char, '_')
            
            # 移除控制字符
            name = ''.join(char for char in name if ord(char) >= 32)
            
            # 限制长度
            if len(name) > 100:
                name = name[:100]
            
            return name.strip()
        
        # 生成文件名
        filename_title = song_name
        song_filename = f"{sanitize_filename(filename_title)}.mp3"
        song_filepath = os.path.join(app.config['SONG_DOWNLOAD_DIR'], song_filename)
        
        # 保存歌曲文件
        with open(song_filepath, 'wb') as f:
            for chunk in song_response.iter_content(chunk_size=8192):
                f.write(chunk)
        
        return song_filepath, "下载成功", lyric
    
    except requests.exceptions.RequestException as e:
        return None, f"网络错误: {str(e)}", None
    except Exception as e:
        return None, f"下载失败: {str(e)}", None

@app.route('/download_song_file/<filename>')
def download_song_file(filename):
    """
    提供歌曲文件下载
    """
    try:
        # 防止路径遍历攻击
        filename = os.path.basename(filename)
        file_path = os.path.join(app.config['SONG_DOWNLOAD_DIR'], filename)
        
        if not os.path.exists(file_path):
            return "文件不存在", 404
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='audio/mpeg'
        )
    except Exception as e:
        return f"下载失败: {str(e)}", 500
def add_song_urls_to_requests(requests_list):
    """为歌曲列表添加播放URL"""
    for request in requests_list:
        song_name = request.get('song_name', '')
        
        def sanitize_filename(name):
            illegal_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
            for char in illegal_chars:
                name = name.replace(char, '_')
            name = ''.join(char for char in name if ord(char) >= 32)
            if len(name) > 100:
                name = name[:100]
            return name.strip()
        
        filename_title = song_name
        expected_filename = f"{sanitize_filename(filename_title)}.mp3"
        expected_filepath = os.path.join(app.config['SONG_DOWNLOAD_DIR'], expected_filename)
        
        if os.path.exists(expected_filepath):
            request['url'] = f"/data/downloads/{expected_filename}"
        else:
            song_id = request.get('song_id')
            if song_id:
                try:
                    params = {
                        'url': song_id,
                        'level': 'standard',
                        'type': 'json'
                    }
                    response = requests.get(
                        'https://api.zh-mc.top/Song_V1',
                        params=params,
                        timeout=10
                    )
                    response.raise_for_status()
                    song_data = response.json()
                    if song_data.get('success') and song_data.get('data'):
                        request['url'] = song_data['data'].get('url', '')
                    else:
                        request['url'] = ''
                except Exception as e:
                    app.logger.error(f"获取歌曲URL失败: {str(e)}")
                    request['url'] = ''
            else:
                request['url'] = ''
    return requests_list

# 修改 add_song_request 函数以添加审核状态
def add_song_request(song_name, class_name, student_name, song_id=None, cover_url=None, artists=None, album=None, lyric=None):
    """添加歌曲请求到当天列表"""
    # 检查点歌是否被暂停
    if is_requests_paused():
        return False, "点歌功能已暂停，请稍后再试"
    
    # 检查是否达到每日限制
    if get_daily_request_count() >= MAX_DAILY_REQUESTS:
        return False, "今日点歌数量已达上限，请明天再来"
    
    # 清理输入数据
    song_name = sanitize_input(song_name, 100)
    class_name = sanitize_input(class_name, 50)
    student_name = sanitize_input(student_name, 50)
    
    # 解析歌曲名称和歌手（如果未提供艺术家信息）
    artist = ""
    title = song_name
    today_list = get_daily_list()
    
    # 检查歌曲是否已存在（基于歌曲ID或歌曲名和歌手）
    for item in today_list:
        if song_id and item.get('song_id') == song_id:
            return False, "该歌曲已被点过，请选择其他歌曲"
        if item['song_name'].lower() == title.lower() and item.get('artist', '').lower() == artist.lower():
            return False, "该歌曲已被点过，请选择其他歌曲"
    
    # 检查同一姓名是否已点过歌
    for item in today_list:
        if item['student_name'] == student_name:
            return False, "您已经点过一首歌了，每人只能点一首"
    
    # 生成唯一ID
    new_id = max([item['id'] for item in today_list], default=0) + 1
    
    # 添加新请求
    new_song_request = {
        'id': new_id,
        'song_name': title,
        'class_name': class_name,
        'student_name': student_name,
        'request_date': datetime.now().isoformat(),
        'votes': 0,
    }
    
    # 如果提供了歌曲ID、封面URL、艺术家、专辑信息和歌词，则添加到记录中
    if song_id:
        new_song_request['song_id'] = song_id
    if cover_url:
        new_song_request['cover_url'] = cover_url
    if artists:
        new_song_request['artists'] = artists
    if album:
        new_song_request['album'] = album
    if lyric:
        new_song_request['lyric'] = lyric
    
    # 确保始终有这些字段
    if 'cover_url' not in new_song_request:
        new_song_request['cover_url'] = ''
    if 'song_id' not in new_song_request:
        new_song_request['song_id'] = ''
    if 'artists' not in new_song_request:
        new_song_request['artists'] = ''
    if 'album' not in new_song_request:
        new_song_request['album'] = ''
    if 'lyric' not in new_song_request:
        new_song_request['lyric'] = ''
    
    today_list.append(new_song_request)
    
    if save_daily_list(today_list):
        return True, "点歌请求已提交成功!"
    else:
        return False, "提交失败，请稍后再试"
        
@app.route('/vote/<int:song_id>', methods=['POST'])
def vote_song(song_id):
    """为歌曲投票"""
    # 检查用户是否已经投过票
    if 'voted_songs' not in session:
        session['voted_songs'] = []
    
    if song_id in session['voted_songs']:
        return jsonify({'success': False, 'message': '您已经投过票了'})
    
    today_list = get_daily_list()
    song_found = False
    
    for song in today_list:
        if song['id'] == song_id:
            song['votes'] = song.get('votes', 0) + 1
            song_found = True
            break
    
    if song_found and save_daily_list(today_list):
        # 记录用户已投票
        session['voted_songs'].append(song_id)
        session.modified = True
        return jsonify({'success': True, 'votes': song['votes']})
    else:
        return jsonify({'success': False, 'message': '投票失败'})
        
# 在 app.py 中找到 search_songs 函数并替换为以下代码

@app.route('/search_songs', methods=['POST'])
def search_songs():
    """搜索歌曲API"""
    song_name = request.form.get('song_name', '').strip()
    
    if not song_name:
        return jsonify({'success': False, 'message': '请输入歌曲名称'})
    
    try:
        params = {
            'keyword': song_name,
            'limit': 10
        }
        
        response = requests.get(
            'https://api.zh-mc.top/Search', 
            params=params,
            timeout=10
        )
        response.raise_for_status()
        
        song_data = response.json()
        if song_data.get('success') and song_data.get('data'):
            return jsonify({
                'success': True, 
                'songs': song_data['data']
            })
        else:
            return jsonify({
                'success': False, 
                'message': '未找到相关歌曲'
            })
    except Exception as e:
        app.logger.error(f"搜索歌曲时发生错误: {str(e)}")
        return jsonify({
            'success': False, 
            'message': '搜索失败，请稍后再试'
        })
        
# 添加新的API端点
@app.route('/api/daily_stats')
def api_daily_stats():
    """获取当日点歌统计信息"""
    count = get_daily_request_count()
    remaining = get_remaining_requests()
    return jsonify({
        'count': count,
        'remaining': remaining,
        'max': MAX_DAILY_REQUESTS
    })

def delete_song_request(request_id):
    """从当天列表中删除歌曲请求"""
    # 验证ID是否为整数
    try:
        request_id = int(request_id)
    except (ValueError, TypeError):
        return False
    
    today_list = get_daily_list()
    today_list = [item for item in today_list if item['id'] != request_id]
    return save_daily_list(today_list)

def get_available_dates():
    """获取所有可用的日期列表（按日期倒序）"""
    files = os.listdir(app.config['DATA_DIR'])
    dates = []
    for f in files:
        if f.endswith('.json') and not f.endswith('admin_accounts.json'):
            date_str = f.replace('.json', '')
            if validate_date_string(date_str):
                dates.append(date_str)
    return sorted(dates, reverse=True)[:100]  # 只保留最近100天

# 管理员账户管理
def get_admin_accounts():
    """获取管理员账户信息"""
    accounts_file = app.config['ADMIN_ACCOUNTS_FILE']
    # 确保文件路径在数据目录内
    if not os.path.abspath(accounts_file).startswith(os.path.abspath(app.config['DATA_DIR'])):
        return []
    
    if os.path.exists(accounts_file):
        try:
            with open(accounts_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return []
    return []

def save_admin_accounts(accounts):
    """保存管理员账户信息"""
    accounts_file = app.config['ADMIN_ACCOUNTS_FILE']
    # 确保文件路径在数据目录内
    if not os.path.abspath(accounts_file).startswith(os.path.abspath(app.config['DATA_DIR'])):
        return False
    
    try:
        with open(accounts_file, 'w', encoding='utf-8') as f:
            json.dump(accounts, f, ensure_ascii=False, indent=2)
        return True
    except:
        return False

def init_admin_account():
    """初始化管理员账户（如果不存在）"""
    accounts = get_admin_accounts()
    if not accounts:
        # 创建默认管理员账户
        default_admin = {
            "username": "admin",
            "password": "admin123",
            "role": "admin",  # 添加角色字段
            "created_at": datetime.now().isoformat()
        }
        accounts.append(default_admin)
        
        # 创建control账户
        control_account = {
            "username": "control",
            "password": "lc2025",
            "role": "control",  # 添加角色字段
            "created_at": datetime.now().isoformat()
        }
        accounts.append(control_account)
        
        save_admin_accounts(accounts)
        print(f"已创建默认管理员账户: {default_admin['username']}/{default_admin['password']}")
        print(f"已创建控制账户: {control_account['username']}/{control_account['password']}")
    else:
        # 检查是否已存在control账户，如果不存在则添加
        control_exists = any(account.get('username') == 'control' for account in accounts)
        if not control_exists:
            control_account = {
                "username": "control",
                "password": "lc2025",
                "role": "control",
                "created_at": datetime.now().isoformat()
            }
            accounts.append(control_account)
            save_admin_accounts(accounts)
            print(f"已创建控制账户: {control_account['username']}/{control_account['password']}")

def authenticate_user(username, password):
    """验证用户凭证"""
    # 清理输入
    username = sanitize_input(username, 50)
    password = sanitize_input(password, 100)
    
    accounts = get_admin_accounts()
    for account in accounts:
        if account['username'] == username and account['password'] == password:
            return account  # 返回整个账户信息，包括角色
    return None

# 登录装饰器
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            flash('请先登录管理员账户', 'danger')
            return redirect(url_for('admin_login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

# 控制账户权限装饰器 - 只允许播放、下载和导出功能
def control_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            flash('请先登录管理员账户', 'danger')
            return redirect(url_for('admin_login', next=request.url))
        
        # 检查用户角色
        username = session.get('admin_username')
        accounts = get_admin_accounts()
        user_account = None
        for account in accounts:
            if account['username'] == username:
                user_account = account
                break
        
        # 如果是control账户，只允许访问特定的路由
        allowed_routes_for_control = [
            'admin',  # 主页（播放功能）
            'serve_download_file',  # 下载文件
            'download_songs',  # 下载歌曲包
            'export_requests',  # 导出Excel
            'admin_logout'  # 登出
        ]
        
        if user_account and user_account.get('role') == 'control':
            # 检查当前请求的路由是否在允许列表中
            if request.endpoint not in allowed_routes_for_control:
                flash('您没有权限访问该功能', 'danger')
                return redirect(url_for('admin'))
        
        return f(*args, **kwargs)
    return decorated_function
    

    
# 在创建app实例后添加
@app.template_filter('format_datetime')
def format_datetime_filter(value):
    """格式化日期时间显示"""
    if not value:
        return ""
    
    try:
        # 解析ISO格式日期
        dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
        # 格式化为年月日时分秒
        return dt.strftime('%Y-%m-%d %H:%M')
    except (ValueError, TypeError):
        return value
        
def init_scheduler():
    scheduler = BackgroundScheduler()
    
    # 设置每天中午12:30执行的任务 - 重置当天歌曲列表
    @scheduler.scheduled_job(CronTrigger(hour=12, minute=30))
    def reset_daily_list():
        # 创建今天的空列表文件
        today_filename = get_today_filename(True)
        if not os.path.exists(today_filename):
            save_daily_list([])
            print(f"Created new daily list for {date.today()}")
    
    # 设置每天下午18:00执行的任务 - 清理旧数据
    @scheduler.scheduled_job(CronTrigger(hour=18, minute=0))
    def cleanup_old_data():
        # 清理100天前的旧数据
        keep_days = 100
        cutoff_date = date.today() - timedelta(days=keep_days)
        
        for filename in os.listdir(app.config['DATA_DIR']):
            if filename.endswith('.json') and not filename.endswith('admin_accounts.json'):
                try:
                    file_date = datetime.strptime(filename.replace('.json', ''), '%Y-%m-%d').date()
                    if file_date < cutoff_date:
                        file_path = os.path.join(app.config['DATA_DIR'], filename)
                        # 确保文件路径在数据目录内
                        if os.path.abspath(file_path).startswith(os.path.abspath(app.config['DATA_DIR'])):
                            os.remove(file_path)
                            print(f"Removed old file: {filename}")
                except ValueError:
                    continue  # 跳过格式不正确的文件名
    

    @scheduler.scheduled_job(CronTrigger(hour=2, minute=0))
    def cleanup_download_files():
        """清理下载的歌曲文件"""
        try:
            if os.path.exists(app.config['SONG_DOWNLOAD_DIR']):
                shutil.rmtree(app.config['SONG_DOWNLOAD_DIR'])
                os.makedirs(app.config['SONG_DOWNLOAD_DIR'], exist_ok=True)
                print("已清理下载的歌曲文件")
        except Exception as e:
            print(f"清理下载文件时出错: {str(e)}")
    
    scheduler.start()

# 修改 / 路由函数，确保在审核前所有歌曲都显示
@app.route('/')
def index():
    form = SongRequestForm()
    announcement = get_announcement()
    system_status = get_system_status()

    
    # 获取今日歌曲列表
    today_requests = get_daily_list()
    
    display_songs = today_requests
    review_completed = False
    
    
    # 按投票数从高到低排序
    display_songs_sorted = sorted(display_songs, key=lambda x: x.get('votes', 0), reverse=True)
    
    # 为每个歌曲添加播放URL
    for request in display_songs_sorted:
        # 如果本地有下载的文件，使用本地文件
        song_name = request.get('song_name', '')
        
        def sanitize_filename(name):
            illegal_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
            for char in illegal_chars:
                name = name.replace(char, '_')
            name = ''.join(char for char in name if ord(char) >= 32)
            if len(name) > 100:
                name = name[:100]
            return name.strip()
        
        filename_title = song_name
        expected_filename = f"{sanitize_filename(filename_title)}.mp3"
        expected_filepath = os.path.join(app.config['SONG_DOWNLOAD_DIR'], expected_filename)
        
        # 检查本地文件是否存在
        if os.path.exists(expected_filepath):
            request['url'] = f"/data/downloads/{expected_filename}"
        else:
            # 如果没有本地文件，但有song_id，则获取在线URL
            song_id = request.get('song_id')
            if song_id:
                try:
                    params = {
                        'url': song_id,
                        'level': 'standard',
                        'type': 'json'
                    }
                    response = requests.get(
                        'https://api.zh-mc.top/Song_V1',
                        params=params,
                        timeout=10
                    )
                    response.raise_for_status()
                    song_data = response.json()
                    if song_data.get('success') and song_data.get('data'):
                        request['url'] = song_data['data'].get('url', '')
                    else:
                        request['url'] = ''
                except Exception as e:
                    app.logger.error(f"获取歌曲URL失败: {str(e)}")
                    request['url'] = ''
            else:
                request['url'] = ''
    
    return render_template('index.html', 
                          form=form, 
                          songs=display_songs_sorted,
                          today_requests_sorted=display_songs_sorted,
                          grade_class_options=json.dumps(GRADE_CLASS_OPTIONS),
                          announcement_enabled=announcement['enabled'],
                          MAX_DAILY_REQUESTS=MAX_DAILY_REQUESTS,
                          system_status=system_status,)


def init_scheduler():
    scheduler = BackgroundScheduler()
    
    # 设置每天中午12:30执行的任务 - 重置当天歌曲列表
    @scheduler.scheduled_job(CronTrigger(hour=12, minute=30))
    def reset_daily_list():
        # 创建今天的空列表文件
        today_filename = get_today_filename(True)
        if not os.path.exists(today_filename):
            save_daily_list([])
            print(f"Created new daily list for {date.today()}")
    
    # 设置每天下午18:00执行的任务 - 清理旧数据
    @scheduler.scheduled_job(CronTrigger(hour=18, minute=0))
    def cleanup_old_data():
        # 清理100天前的旧数据
        keep_days = 100
        cutoff_date = date.today() - timedelta(days=keep_days)
        
        for filename in os.listdir(app.config['DATA_DIR']):
            if filename.endswith('.json') and not filename.endswith('admin_accounts.json') and not filename.endswith('review_status.json'):
                try:
                    file_date = datetime.strptime(filename.replace('.json', ''), '%Y-%m-%d').date()
                    if file_date < cutoff_date:
                        file_path = os.path.join(app.config['DATA_DIR'], filename)
                        # 确保文件路径在数据目录内
                        if os.path.abspath(file_path).startswith(os.path.abspath(app.config['DATA_DIR'])):
                            os.remove(file_path)
                            print(f"Removed old file: {filename}")
                except ValueError:
                    continue  # 跳过格式不正确的文件名
    
    # 添加每日清理下载文件的任务（凌晨2点执行）
    @scheduler.scheduled_job(CronTrigger(hour=2, minute=0))
    def cleanup_download_files():
        """清理下载的歌曲文件"""
        try:
            if os.path.exists(app.config['SONG_DOWNLOAD_DIR']):
                shutil.rmtree(app.config['SONG_DOWNLOAD_DIR'])
                os.makedirs(app.config['SONG_DOWNLOAD_DIR'], exist_ok=True)
                print("已清理下载的歌曲文件")
        except Exception as e:
            print(f"清理下载文件时出错: {str(e)}")
    
    scheduler.start()

@app.route('/get_classes/<grade>')
def get_classes(grade):
    """获取对应年级的班级列表"""
    if grade in GRADE_CLASS_OPTIONS:
        return jsonify(GRADE_CLASS_OPTIONS[grade])
    return jsonify([])
    
# 在 app.py 中添加读取更新日志的函数
def get_changelog():
    """获取更新日志内容"""
    changelog_file = os.path.join(app.config['DATA_DIR'], 'changelog.json')
    # 确保文件路径在数据目录内
    if not os.path.abspath(changelog_file).startswith(os.path.abspath(app.config['DATA_DIR'])):
        return []
    
    if os.path.exists(changelog_file):
        try:
            with open(changelog_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return []
    return []

# 修改更新日志路由
@app.route('/changelog')
def changelog():
    """更新日志页面"""
    changelog_data = get_changelog()
    return render_template('changelog.html', changelog=changelog_data)



@app.route('/submit', methods=['POST'])
def submit_request():
    form = SongRequestForm()
    system_status = get_system_status()
    
    # 动态设置班级选项
    if form.grade.data:
        form.class_name.choices = [(cls, cls) for cls in GRADE_CLASS_OPTIONS.get(form.grade.data, [])]
    else:
        form.class_name.choices = [('', '请先选择年段')]
    
    if form.validate_on_submit():
        # 获取歌曲相关信息
        song_id = request.form.get('song_id')
        cover_url = request.form.get('cover_url')
        artists = request.form.get('artists')
        album = request.form.get('album')
        
        # 如果有歌曲ID，获取歌词信息
        lyric = None
        if song_id:
            try:
                params = {
                    'url': song_id,
                    'level': 'standard',
                    'type': 'json'
                }
                
                response = requests.get(
                    'https://api.zh-mc.top/Song_V1',
                    params=params,
                    timeout=10
                )
                response.raise_for_status()
                
                song_data = response.json()
                if song_data.get('success') and song_data.get('data'):
                    lyric = song_data['data'].get('lyric', '')
            except Exception as e:
                app.logger.error(f"获取歌词时发生错误: {str(e)}")
        
        success, message = add_song_request(
            form.song_name.data,
            form.class_name.data,
            form.student_name.data,
            song_id if song_id else None,
            cover_url if cover_url else None,
            artists if artists else None,
            album if album else None,
            lyric if lyric else None
        )
        
        if success:
            flash(message, 'success')
            # 如果有歌曲ID，准备下载信息并触发下载
            if song_id:
                # 直接在后台下载歌曲
                download_file_path, download_message, _ = download_single_song(
                    song_id, 
                    form.song_name.data, 
                    artists if artists else ''
                )
                
                # 将下载链接存储在session中，供前端获取
                if download_file_path:
                    filename = os.path.basename(download_file_path)
                    session['download_url'] = url_for('download_song_file', filename=filename)
                else:
                    flash(f'歌曲下载失败: {download_message}', 'warning')
            
            return redirect(url_for('index'))
        else:
            flash(message, 'danger')
            # 获取今日歌曲列表并按投票数排序，确保模板变量完整
            today_requests = get_daily_list()
            today_requests_sorted = sorted(today_requests, key=lambda x: x.get('votes', 0), reverse=True)
            today_requests_sorted = add_song_urls_to_requests(today_requests_sorted)
            
            return render_template('index.html', 
                                form=form, 
                                songs=today_requests_sorted,
                                today_requests_sorted=today_requests_sorted,
                                grade_class_options=json.dumps(GRADE_CLASS_OPTIONS),
                                system_status=system_status,
                                announcement_enabled=get_announcement()['enabled'],
                                MAX_DAILY_REQUESTS=MAX_DAILY_REQUESTS)
    
    # 如果表单验证失败，返回错误信息
    for field, errors in form.errors.items():
        for error in errors:
            flash(f"{getattr(form, field).label.text}: {error}", 'danger')
    
    # 获取今日歌曲列表并按投票数排序，确保模板变量完整
    today_requests = get_daily_list()
    today_requests_sorted = sorted(today_requests, key=lambda x: x.get('votes', 0), reverse=True)
    today_requests_sorted = add_song_urls_to_requests(today_requests_sorted)
    
    return render_template('index.html',
                          form=form,
                          songs=today_requests_sorted,
                          today_requests_sorted=today_requests_sorted,
                          grade_class_options=json.dumps(GRADE_CLASS_OPTIONS),
                          system_status=system_status,
                          announcement_enabled=get_announcement()['enabled'],
                          MAX_DAILY_REQUESTS=MAX_DAILY_REQUESTS)
@app.route('/contact')
def contact():
    """联系我们页面"""
    return render_template('contact.html')
    
# 管理员登录路由
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    # 如果已登录，重定向到管理页面
    if session.get('admin_logged_in'):
        return redirect(url_for('admin'))
    
    form = LoginForm()
    if form.validate_on_submit():
        username = form.username.data
        password = form.password.data
        
        account = authenticate_user(username, password)
        if account:
            session['admin_logged_in'] = True
            session['admin_username'] = username
            session['admin_role'] = account.get('role', 'admin')  # 保存角色信息
            flash('登录成功!', 'success')
            
            # 重定向到请求的页面或管理主页
            next_page = request.args.get('next')
            # 验证next参数是否安全（防止开放重定向）
            if next_page and not next_page.startswith('//') and not next_page.startswith('http'):
                return redirect(next_page)
            return redirect(url_for('admin'))
        else:
            flash('用户名或密码错误', 'danger')
    
    return render_template('admin_login.html', form=form)
# 管理员登出
@app.route('/admin/logout')
@admin_required
def admin_logout():
    session.pop('admin_logged_in', None)
    session.pop('admin_username', None)
    flash('您已成功登出', 'success')
    return redirect(url_for('index'))

@app.route('/data/downloads/<path:filename>')
def serve_download_file(filename):
    """提供对下载目录中文件的直接访问 - 跨平台版本"""
    try:
        # 解码URL编码的文件名
        filename = urllib.parse.unquote(filename)
        
        # 确保只取文件名部分，防止路径遍历
        filename = os.path.basename(filename)
        
        # 构建文件路径
        file_path = os.path.join(app.config['SONG_DOWNLOAD_DIR'], filename)
        
        # 检查文件是否存在
        if not os.path.exists(file_path):
            app.logger.error(f"文件不存在: {file_path}")
            # 列出目录中的所有文件以帮助调试
            if os.path.exists(app.config['SONG_DOWNLOAD_DIR']):
                files = os.listdir(app.config['SONG_DOWNLOAD_DIR'])
                app.logger.error(f"目录中的文件: {files}")
            return "文件不存在", 404
            
        return send_file(
            file_path,
            mimetype='audio/mpeg'
        )
    except Exception as e:
        app.logger.error(f"提供下载文件时出错: {str(e)}")
        return "文件访问失败", 500
@app.route('/admin')
@control_required
def admin():
    # 获取今日歌单
    today_requests = get_daily_list()
    # 按投票数从高到低排序
    today_requests_sorted = sorted(today_requests, key=lambda x: x.get('votes', 0), reverse=True)
    system_status = get_system_status()
    
    # 获取当前用户角色
    current_user_role = session.get('admin_role', 'admin')
    
    # 为每个歌曲添加本地MP3文件的URL
    for request in today_requests_sorted:
        song_name = request.get('song_name', '')
        # 清理文件名中的非法字符
        def sanitize_filename(name):
            """统一的文件名清理函数"""
            # 移除或替换非法字符
            illegal_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
            for char in illegal_chars:
                name = name.replace(char, '_')
            
            # 移除控制字符
            name = ''.join(char for char in name if ord(char) >= 32)
            
            # 限制长度
            if len(name) > 100:
                name = name[:100]
            
            return name.strip()
        
        # 构造可能已存在的文件名
        filename_title = song_name
        expected_filename = f"{sanitize_filename(filename_title)}.mp3"
        expected_filepath = os.path.join(app.config['SONG_DOWNLOAD_DIR'], expected_filename)
        
        # 检查文件是否存在，如果存在则添加URL
        if os.path.exists(expected_filepath):
            # 直接使用相对路径，避免URL编码问题
            request['url'] = f"/data/downloads/{expected_filename}"
        else:
            # 如果本地文件不存在，但有song_id，则尝试获取在线URL
            song_id = request.get('song_id')
            if song_id:
                try:
                    params = {
                        'url': song_id,
                        'level': 'standard',
                        'type': 'json'
                    }
                    response = requests.get(
                        'https://api.zh-mc.top/Song_V1',
                        params=params,
                        timeout=10
                    )
                    response.raise_for_status()
                    song_data = response.json()
                    if song_data.get('success') and song_data.get('data'):
                        request['url'] = song_data['data'].get('url', '')
                        # 同时获取歌词
                        request['lyric'] = song_data['data'].get('lyric', '')
                    else:
                        request['url'] = ''
                        request['lyric'] = ''
                except Exception as e:
                    app.logger.error(f"获取歌曲URL失败: {str(e)}")
                    request['url'] = ''
                    request['lyric'] = ''
            else:
                request['url'] = ''
                request['lyric'] = ''
    
    return render_template('admin.html', 
                          today_requests=today_requests_sorted,
                          system_status=system_status,
                          user_role=current_user_role)  # 传递用户角色到模板
# 批量删除歌曲请求
@app.route('/admin/batch_delete', methods=['POST'])
@admin_required
def batch_delete():
    """批量删除歌曲请求"""
    try:
        # 获取选中的歌曲ID列表
        selected_songs_json = request.form.get('selected_songs')
        if not selected_songs_json:
            flash('未选择任何歌曲', 'warning')
            return redirect(url_for('admin'))
        
        selected_songs = json.loads(selected_songs_json)
        if not selected_songs:
            flash('未选择任何歌曲', 'warning')
            return redirect(url_for('admin'))
        
        # 转换为整数ID
        selected_ids = [int(song_id) for song_id in selected_songs]
        
        # 获取当前歌曲列表
        today_list = get_daily_list()
        
        # 过滤掉被选中的歌曲
        filtered_list = [song for song in today_list if song['id'] not in selected_ids]
        
        # 保存更新后的列表
        if save_daily_list(filtered_list):
            flash(f'成功删除 {len(today_list) - len(filtered_list)} 首歌曲', 'success')
        else:
            flash('删除失败，请稍后再试', 'danger')
            
    except Exception as e:
        flash('删除过程中发生错误', 'danger')
        app.logger.error(f'批量删除歌曲时发生错误: {str(e)}')
    
    return redirect(url_for('admin'))
@app.route('/admin/export')
@admin_required
def export_requests():
    """导出今日点歌列表为Excel文件"""
    # 获取今日点歌列表
    today_requests = get_daily_list()
    
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "今日点歌列表"
    
    # 设置标题行样式
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")
    
    # 写入标题行
    headers = ['ID', '歌曲名称', '班级', '姓名', '点歌时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 写入数据行
    for row, request in enumerate(today_requests, 2):
        ws.cell(row=row, column=1, value=request['id'])
        ws.cell(row=row, column=2, value=request['song_name'])
        ws.cell(row=row, column=3, value=request['class_name'])
        ws.cell(row=row, column=4, value=request['student_name'])
        
        # 格式化时间
        request_time = request.get('request_date', '')
        if request_time:
            try:
                dt = datetime.fromisoformat(request_time)
                request_time = dt.strftime('%Y-%m-%d %H:%M:%S')
            except:
                pass
        ws.cell(row=row, column=5, value=request_time)
    
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 将Excel文件保存到内存中
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 发送文件作为响应
    today_str = date.today().strftime('%Y-%m-%d')
    filename = f'点歌列表_{today_str}.xlsx'
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# 在 app.py 中添加自动审核相关函数和路由

def auto_review_songs(songs_list):
    """
    使用 DeepSeek API 自动审核歌曲列表
    """
    try:
        # 将歌曲列表转换为 JSON 字符串
        songs_json = json.dumps(songs_list, ensure_ascii=False, indent=2)
        
        # 构建提示词
        prompt = f"""
你是一个校园点歌台的审核员
请根据以下规则，对以下歌曲进行审核：

审核规则：
1. 禁止包含暴力、色情、低俗内容的歌曲
2. 禁止包含不健康、不积极的歌曲
3. 禁止过于吵闹或不适合校园环境的歌曲
4. 优先通过积极向上、旋律优美的歌曲
6. 不允许日语的歌曲
7. 符合社会主义核心价值观


请审核以下歌曲列表，并以指定的JSON格式输出结果：

歌曲列表：
{songs_json}

请按以下格式输出审核结果：
[
  {{
    "歌曲名称": "歌曲名",
    "是否通过": true/false,
    "原因": "通过原因或拒绝理由"
  }},
  ...
]

请只输出JSON结果，不要包含其他内容。
"""
        
        # 初始化 DeepSeek 客户端
        client = OpenAI(
            api_key=app.config['DEEPSEEK_API_KEY'],
            base_url="https://api.deepseek.com/v1"
        )
        
        # 调用 API 进行审核
        response = client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": "你是一个严格的校园点歌台审核员，负责审核学生点播的歌曲是否适合在校园播放。"},
                {"role": "user", "content": prompt}
            ],
            stream=False,
            temperature=0.3  # 降低随机性，使结果更稳定
        )
        
        # 获取审核结果
        review_result = response.choices[0].message.content
        
        # 尝试解析 JSON 结果
        print(review_result)
        
        # 提取JSON部分（去除可能的代码块标记）
        if review_result.startswith("```json"):
            review_result = review_result[7:]  # 移除开头的 ```json
        if review_result.endswith("```"):
            review_result = review_result[:-3]  # 移除结尾的 ```
        
        # 解析JSON
        result_json = json.loads(review_result)
        return result_json
            
    except json.JSONDecodeError as e:
        app.logger.error(f"JSON解析错误: {e}")
        app.logger.error(f"原始响应内容: {review_result}")
        return None
    except Exception as e:
        app.logger.error(f"调用DeepSeek API时出错: {e}")
        return None

# 添加全局变量存储最近一次的审核结果
recent_review_results = []

@app.route('/admin/auto_review', methods=['POST'])
@admin_required
def admin_auto_review():
    """
    自动审核歌曲列表
    """
    try:
        # 获取今日歌曲列表
        today_requests = get_daily_list()
        
        if not today_requests:
            return jsonify({
                'status': 'error',
                'message': '今日无点歌记录'
            }), 400
        
        # 调用自动审核函数
        review_results = auto_review_songs(today_requests)
        
        if review_results is None:
            return jsonify({
                'status': 'error',
                'message': '审核失败，请稍后重试'
            }), 500
        
        # 存储审核结果以便后续应用
        global recent_review_results
        recent_review_results = review_results
        
        return jsonify({
            'status': 'success',
            'results': review_results
        })
        
    except Exception as e:
        app.logger.error(f"自动审核出错: {e}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500



@app.route('/admin/apply_review_results', methods=['POST'])
@admin_required
def apply_review_results():
    """
    应用审核结果到歌曲列表
    """
    try:
        global recent_review_results
        
        # 获取请求数据
        data = request.get_json()
        selected_indices = data.get('indices', [])
        
        if not selected_indices:
            return jsonify({
                'status': 'error',
                'message': '未选择任何审核结果'
            }), 400
        
        # 获取当前歌曲列表
        today_requests = get_daily_list()
        
        # 解析 recent_review_results（可能是字符串或已解析的列表）
        review_results = recent_review_results
        if isinstance(review_results, str):
            try:
                # 如果是字符串，尝试解析为JSON
                review_results = json.loads(review_results)
            except json.JSONDecodeError as e:
                app.logger.error(f"审核结果JSON解析错误: {e}")
                return jsonify({
                    'status': 'error',
                    'message': '审核结果格式错误：无法解析JSON'
                }), 500
        
        # 确保 review_results 是一个列表
        if not isinstance(review_results, list):
            app.logger.error(f"审核结果不是列表类型: {type(review_results)}")
            return jsonify({
                'status': 'error',
                'message': '审核结果格式错误：不是有效的列表格式'
            }), 500
        
        # 创建一个集合来存储通过审核的歌曲名称
        approved_songs = set()
        applied_count = 0
        
        # 收集通过审核的歌曲名称
        for index in selected_indices:
            if 0 <= index < len(review_results):
                review_result = review_results[index]
                # 确保 review_result 是一个字典
                if isinstance(review_result, dict):
                    song_name = review_result.get('歌曲名称')
                    passed = review_result.get('是否通过', True)
                    
                    # 如果歌曲通过审核，则添加到批准列表
                    if passed and song_name:
                        approved_songs.add(song_name)
                    
                    app.logger.info(f"应用审核结果: {song_name} - {'通过' if passed else '不通过'} - {review_result.get('原因', '无原因')}")
                    applied_count += 1
                else:
                    app.logger.error(f"审核结果项不是字典类型: {type(review_result)}")
        
        # 过滤掉未通过审核的歌曲（保留在approved_songs中的歌曲）
        filtered_list = [song for song in today_requests if song.get('song_name') in approved_songs]
        deleted_count = len(today_requests) - len(filtered_list)
        
        # 保存更新后的列表
        if save_daily_list(filtered_list):
            app.logger.info(f"已删除 {deleted_count} 首未通过审核的歌曲")
        else:
            app.logger.error("保存更新后的歌曲列表失败")
            return jsonify({
                'status': 'error',
                'message': '保存歌曲列表失败'
            }), 500
        
        return jsonify({
            'status': 'success',
            'message': f'成功应用 {applied_count} 条审核结果，删除了 {deleted_count} 首歌曲'
        })
        
    except Exception as e:
        app.logger.error(f"应用审核结果出错: {e}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/admin/download_songs')
@admin_required
def download_songs():
    """下载今日点歌列表中的歌曲并打包为ZIP"""
   
    # 获取今日点歌列表
    today_requests = get_daily_list()
    if not today_requests:
        app.logger.warning("今日无点歌记录")
        return jsonify({'status': 'error', 'message': '今日无点歌记录'}), 400
        
    # 按投票数从高到低排序
    today_requests_sorted = sorted(today_requests, key=lambda x: x.get('votes', 0), reverse=True)
        
    app.logger.info(f"开始下载 {len(today_requests_sorted)} 首歌曲")
        
    # 创建临时目录存放下载的歌曲
    temp_dir = tempfile.mkdtemp()
    zip_filename = f"songs_{date.today().isoformat()}.zip"
    zip_path = os.path.join(tempfile.gettempdir(), zip_filename)
        
    # 记录成功和失败的歌曲数量
    success_count = 0
    error_count = 0
    
    try:
        # 创建ZIP文件
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            # 下载每首歌曲并添加到ZIP
            for index, request in enumerate(today_requests_sorted, 1):
                song_name = request['song_name']
                artist = request.get('artist', '')
                votes = request.get('votes', 0)
                song_id = request.get('song_id')  # 获取歌曲ID
                
                # 如果没有歌曲ID，跳过下载
                if not song_id:
                    app.logger.warning(f"歌曲 '{song_name}' 缺少ID，跳过下载")
                    error_count += 1
                    continue
                
                try:
                    # 清理文件名中的非法字符
                    def sanitize_filename(name):
                        """统一的文件名清理函数"""
                        # 移除或替换非法字符
                        illegal_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
                        for char in illegal_chars:
                            name = name.replace(char, '_')
                        
                        # 移除控制字符
                        name = ''.join(char for char in name if ord(char) >= 32)
                        
                        # 限制长度
                        if len(name) > 100:
                            name = name[:100]
                        
                        return name.strip()     
                    
                    # 构造可能已存在的文件名

                    filename_title = song_name
                    expected_filename = f"{sanitize_filename(filename_title)}.mp3"
                    expected_filepath = os.path.join(app.config['SONG_DOWNLOAD_DIR'], expected_filename)
                    
                    # 检查文件是否已存在
                    if os.path.exists(expected_filepath):
                        # 使用已存在的文件
                        song_filepath = expected_filepath
                        app.logger.info(f"使用已存在的文件: {expected_filename}")
                    else:
                        # 文件不存在，需要下载
                        song_filepath, message, lyric = download_single_song(song_id, song_name, artist)  # 修复这一行
                        if not song_filepath:
                            app.logger.error(f"下载歌曲失败: {song_name}, 错误: {message}")
                            error_count += 1
                            continue
                    
                    # 重新命名文件以包含序号和投票数
                    new_filename = f"{index:02d}_{votes}票_{sanitize_filename(filename_title)}.mp3"
                    new_filepath = os.path.join(temp_dir, new_filename)
                    
                    # 复制文件到临时目录并重命名
                    shutil.copy2(song_filepath, new_filepath)
                    
                    # 添加到ZIP文件
                    zipf.write(new_filepath, arcname=new_filename)
                    success_count += 1
                    
                except Exception as e:
                    app.logger.error(f"处理歌曲 '{song_name}' 时发生错误: {str(e)}")
                    error_count += 1
                    continue
        
        # 返回结果，包含成功和失败的统计信息
        return jsonify({
            'status': 'success',
            'message': f'下载完成: 成功 {success_count} 首，失败 {error_count} 首',
            'success_count': success_count,
            'error_count': error_count,
            'download_url': url_for('download_zip', filename=zip_filename)
        })
    
    except Exception as e:
        app.logger.error(f"下载歌曲时发生未知错误: {str(e)}")
        return jsonify({'status': 'error', 'message': f'系统错误: {str(e)}'}), 500
    
    finally:
        # 清理临时文件
        def safe_delete(path):
            try:
                if os.path.isfile(path):
                    os.remove(path)
                elif os.path.isdir(path):
                    shutil.rmtree(path)
            except Exception as e:
                app.logger.warning(f"删除临时文件失败: {path}, 错误: {str(e)}")
        
        safe_delete(temp_dir)
@app.route('/download_zip/<filename>')
def download_zip(filename):
    """提供ZIP文件下载"""
    zip_path = os.path.join(tempfile.gettempdir(), filename)
    
    if not os.path.exists(zip_path):
        return "文件不存在或已过期", 404
    
    return send_file(
        zip_path,
        as_attachment=True,
        download_name=filename,
        mimetype='application/zip'
    )

# 删除请求
@app.route('/admin/delete/<int:request_id>', methods=['POST'])
@admin_required
def delete_request(request_id):
    if delete_song_request(request_id):
        flash('点歌请求已删除!', 'success')
    else:
        flash('删除失败，请稍后再试', 'danger')
    return redirect(url_for('admin'))

# 暂停/恢复点歌功能
@app.route('/admin/toggle_pause', methods=['POST'])
@admin_required
def toggle_pause():
    """切换点歌暂停状态"""
    status = get_system_status()
    status['requests_paused'] = not status.get('requests_paused', False)
    
    if status['requests_paused']:
        status['pause_reason'] = request.form.get('reason', '点歌功能已暂停')
    else:
        status['pause_reason'] = ''
    
    if save_system_status(status):
        if status['requests_paused']:
            flash('点歌功能已暂停', 'success')
        else:
            flash('点歌功能已恢复', 'success')
    else:
        flash('操作失败，请稍后再试', 'danger')
    
    return redirect(url_for('admin'))

# 清空歌曲列表
@app.route('/admin/clear_list', methods=['POST'])
@admin_required
def clear_list():
    """清空今日点歌列表"""
    # 验证管理员密码
    password = request.form.get('password', '')
    username = session.get('admin_username', 'admin')
    
    if authenticate_user(username, password):
        today_list = get_daily_list()
        if not today_list:
            flash('今日点歌列表已为空', 'warning')
        else:
            # 清空列表
            if save_daily_list([]):
                flash('今日点歌列表已清空', 'success')
            else:
                flash('清空失败，请稍后再试', 'danger')
    else:
        flash('密码错误，操作失败', 'danger')
    
    return redirect(url_for('admin'))

# 公告文件路径
app.config['ANNOUNCEMENT_FILE'] = os.path.join(app.config['DATA_DIR'], 'announcement.json')

def get_announcement():
    """获取公告内容"""
    announcement_file = app.config['ANNOUNCEMENT_FILE']
    # 确保文件路径在数据目录内
    if not os.path.abspath(announcement_file).startswith(os.path.abspath(app.config['DATA_DIR'])):
        return {"content": "", "enabled": False}
    
    if os.path.exists(announcement_file):
        try:
            with open(announcement_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {"content": "", "enabled": False}
    return {"content": "", "enabled": False}

def save_announcement(content, enabled):
    """保存公告内容"""
    announcement_file = app.config['ANNOUNCEMENT_FILE']
    # 确保文件路径在数据目录内
    if not os.path.abspath(announcement_file).startswith(os.path.abspath(app.config['DATA_DIR'])):
        return False
    
    try:
        with open(announcement_file, 'w', encoding='utf-8') as f:
            json.dump({"content": content, "enabled": enabled}, f, ensure_ascii=False, indent=2)
        return True
    except:
        return False

# 公告管理路由
@app.route('/admin/announcement', methods=['GET', 'POST'])
@admin_required
def admin_announcement():
    announcement = get_announcement()
    
    if request.method == 'POST':
        content = request.form.get('content', '')
        enabled = 'enabled' in request.form
        
        # 清理输入
        # content = sanitize_input(content, 1000)  # 限制公告长度
        
        if save_announcement(content, enabled):
            flash('公告更新成功!', 'success')
        else:
            flash('公告更新失败!', 'danger')
        
        return redirect(url_for('admin_announcement'))
    
    return render_template('admin_announcement.html', 
                          announcement_content=announcement['content'],
                          announcement_enabled=announcement['enabled'])

# 获取公告API
@app.route('/api/announcement')
def api_announcement():
    announcement = get_announcement()
    return jsonify(announcement)

# 在模板中获取系统状态
@app.template_global()
def get_system_status_global():
    """在模板中获取系统状态"""
    return get_system_status()

# 应用初始化
if __name__ == '__main__':
    # 初始化管理员账户
    init_admin_account()
    
    # 启动定时任务
    init_scheduler()
    
    app.run(debug=True)